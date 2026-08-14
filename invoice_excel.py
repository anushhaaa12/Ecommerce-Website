"""
invoice_excel.py
-----------------
Writes extraction results (one or more invoices) to a single Excel workbook
with 4 sheets: Classification Summary, Header & Parties, Line Items,
Additional Fields. Applies formatting per spec: bold headers, frozen header
row, auto column width, highlighted null/unclear cells.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
NULL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
UNCLEAR_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # light red
ERROR_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")


def _style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)


def _autosize(ws, ncols, max_width=60):
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for cell in ws[col_letter]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _write_value_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if value is None:
        cell.value = "null"
        cell.fill = NULL_FILL
        cell.font = Font(name=FONT_NAME, italic=True, color="7F6000")
    elif isinstance(value, str) and value.strip() == "[unclear]":
        cell.value = "[unclear]"
        cell.fill = UNCLEAR_FILL
        cell.font = Font(name=FONT_NAME, italic=True, color="9C0006")
    else:
        cell.value = value
        cell.font = Font(name=FONT_NAME)
    return cell


def write_to_excel(results, output_path):
    """
    results: list of dicts as returned by invoice_extract.process_invoice(),
             or a dict with status='Error' for failed files
             (must include filename, status, error keys at minimum).
    """
    wb = Workbook()

    # ---------------- Sheet 1: Classification Summary ----------------
    ws1 = wb.active
    ws1.title = "Classification Summary"
    headers1 = ["Filename", "Status", "Invoice Type", "Type Confidence (%)",
                "Type Evidence", "Country", "Country Evidence", "Memo Flag",
                "Page Count", "OCR Pages", "Validation Notes", "Error"]
    for c, h in enumerate(headers1, start=1):
        ws1.cell(row=1, column=c, value=h)
    _style_header_row(ws1, 1, len(headers1))

    r = 2
    for res in results:
        if res.get("status") == "Error":
            row_vals = [res.get("filename"), "Error", None, None, None, None,
                        None, None, None, None, None, res.get("error")]
            for c, v in enumerate(row_vals, start=1):
                cell = _write_value_cell(ws1, r, c, v)
                if c == 2:
                    cell.fill = ERROR_FILL
            r += 1
            continue

        cl = res["classification"]
        row_vals = [
            res["filename"], "OK", cl["invoice_type"], cl["type_confidence"],
            "; ".join(cl["type_evidence"]) if cl["type_evidence"] else None,
            cl["country"],
            "; ".join(cl["country_evidence"]) if cl["country_evidence"] else None,
            cl["memo_flag"], res["page_count"],
            ", ".join(str(p) for p in res["ocr_pages"]) if res["ocr_pages"] else None,
            "; ".join(res["validation_notes"]), None,
        ]
        for c, v in enumerate(row_vals, start=1):
            _write_value_cell(ws1, r, c, v)
        r += 1
    _autosize(ws1, len(headers1))

    # ---------------- Sheet 2: Header & Parties ----------------
    ws2 = wb.create_sheet("Header & Parties")
    headers2 = ["Filename", "Section", "Field", "Value"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header_row(ws2, 1, len(headers2))

    r = 2
    for res in results:
        if res.get("status") == "Error":
            continue
        for section_key, section_label in [
            ("Header", "Invoice Header"), ("BillTo", "Bill To Details"),
            ("ShipTo", "Ship To Details"), ("Terms", "Terms / Shipment"),
        ]:
            for field, value in res["labeled_fields"].get(section_key, {}).items():
                ws2.cell(row=r, column=1, value=res["filename"]).font = Font(name=FONT_NAME)
                ws2.cell(row=r, column=2, value=section_label).font = Font(name=FONT_NAME)
                ws2.cell(row=r, column=3, value=field).font = Font(name=FONT_NAME)
                _write_value_cell(ws2, r, 4, value)
                r += 1
    _autosize(ws2, len(headers2))

    # ---------------- Sheet 3: Line Items ----------------
    ws3 = wb.create_sheet("Line Items")
    from invoice_extract import CANON_COLUMNS
    headers3 = ["Filename", "Source Page"] + CANON_COLUMNS + ["Extra Columns"]
    for c, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=c, value=h)
    _style_header_row(ws3, 1, len(headers3))

    r = 2
    for res in results:
        if res.get("status") == "Error":
            continue
        for item in res["line_items"]:
            ws3.cell(row=r, column=1, value=res["filename"]).font = Font(name=FONT_NAME)
            ws3.cell(row=r, column=2, value=item.get("_source_page")).font = Font(name=FONT_NAME)
            for ci, col in enumerate(CANON_COLUMNS, start=3):
                _write_value_cell(ws3, r, ci, item.get(col))
            extra = item.get("_extra_columns")
            extra_str = "; ".join(f"{k}={v}" for k, v in extra.items()) if extra else None
            ws3.cell(row=r, column=3 + len(CANON_COLUMNS), value=extra_str).font = Font(name=FONT_NAME)
            r += 1
    _autosize(ws3, len(headers3))

    # ---------------- Sheet 4: Additional Fields ----------------
    ws4 = wb.create_sheet("Additional Fields")
    headers4 = ["Filename", "Label", "Value"]
    for c, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=c, value=h)
    _style_header_row(ws4, 1, len(headers4))

    r = 2
    for res in results:
        if res.get("status") == "Error":
            continue
        for label, value in res.get("additional_fields", []):
            ws4.cell(row=r, column=1, value=res["filename"]).font = Font(name=FONT_NAME)
            ws4.cell(row=r, column=2, value=label).font = Font(name=FONT_NAME)
            _write_value_cell(ws4, r, 3, value)
            r += 1
    _autosize(ws4, len(headers4))

    wb.save(output_path)
    return output_path
